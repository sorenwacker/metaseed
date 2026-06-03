"""Tests for metaseed.cli.commands.example."""

import json

import pytest
import yaml
from typer.testing import CliRunner

from metaseed.cli.commands.example import _export_example_to_excel, export_example

runner = CliRunner()


@pytest.fixture
def miappe_example_data() -> dict:
    """Sample MIAPPE example data structure."""
    return {
        "unique_id": "INV001",
        "title": "Test Investigation",
        "studies": [
            {
                "unique_id": "STU001",
                "title": "Test Study",
                "persons": [{"name": "John Doe", "email": "john@example.com"}],
                "biological_materials": [{"unique_id": "BM001", "genus": "Triticum"}],
                "factors": [
                    {
                        "unique_id": "FAC001",
                        "name": "Temperature",
                        "values": [{"value": "25C"}],
                    }
                ],
                "observed_variables": [{"unique_id": "OV001", "name": "Height"}],
                "observation_units": [
                    {
                        "unique_id": "OU001",
                        "type": "Plot",
                        "samples": [{"unique_id": "SAM001"}],
                        "factor_values": [{"value": "25C"}],
                    }
                ],
                "events": [{"unique_id": "EVT001", "type": "Sowing"}],
                "environments": [
                    {"unique_id": "ENV001", "parameter": "Air temperature"}
                ],
                "data_files": [{"unique_id": "DF001", "link": "data.csv"}],
                "protocols": [{"unique_id": "PROT001", "name": "Sampling"}],
                "sources": [{"unique_id": "SRC001", "name": "Field"}],
                "samples": [{"unique_id": "SAM002"}],
                "assays": [{"unique_id": "ASS001", "measurement_type": "phenotyping"}],
            }
        ],
        "contacts": [{"name": "PI Name", "role": "PI"}],
    }


class TestExportExampleToExcel:
    """Test _export_example_to_excel function."""

    @pytest.mark.skipif(
        not pytest.importorskip("openpyxl", reason="openpyxl not installed"),
        reason="openpyxl required",
    )
    def test_creates_excel_file(self, tmp_path, miappe_example_data):
        """Creates an Excel file with multiple sheets."""
        output_path = tmp_path / "test.xlsx"
        _export_example_to_excel(miappe_example_data, output_path)

        assert output_path.exists()

    @pytest.mark.skipif(
        not pytest.importorskip("openpyxl", reason="openpyxl not installed"),
        reason="openpyxl required",
    )
    def test_creates_investigation_sheet(self, tmp_path, miappe_example_data):
        """Creates Investigation sheet with root entity data."""
        from openpyxl import load_workbook

        output_path = tmp_path / "test.xlsx"
        _export_example_to_excel(miappe_example_data, output_path)

        wb = load_workbook(output_path)
        assert "Investigation" in wb.sheetnames

    @pytest.mark.skipif(
        not pytest.importorskip("openpyxl", reason="openpyxl not installed"),
        reason="openpyxl required",
    )
    def test_creates_study_sheet(self, tmp_path, miappe_example_data):
        """Creates Study sheet when studies present."""
        from openpyxl import load_workbook

        output_path = tmp_path / "test.xlsx"
        _export_example_to_excel(miappe_example_data, output_path)

        wb = load_workbook(output_path)
        assert "Study" in wb.sheetnames

    @pytest.mark.skipif(
        not pytest.importorskip("openpyxl", reason="openpyxl not installed"),
        reason="openpyxl required",
    )
    def test_creates_nested_entity_sheets(self, tmp_path, miappe_example_data):
        """Creates sheets for nested entities within studies."""
        from openpyxl import load_workbook

        output_path = tmp_path / "test.xlsx"
        _export_example_to_excel(miappe_example_data, output_path)

        wb = load_workbook(output_path)
        # At minimum should have Investigation and Study
        assert len(wb.sheetnames) >= 2


class TestExportExampleCommand:
    """Test export_example CLI command using Typer testing."""

    def test_list_examples(self):
        """Lists available examples when called with --list."""
        import typer

        app = typer.Typer()
        app.command()(export_example)

        result = runner.invoke(app, ["--list"])

        # Should not error
        assert result.exit_code == 0
        assert "Available example datasets" in result.stdout

    def test_list_examples_without_profile(self):
        """Lists examples when called without arguments."""
        import typer

        app = typer.Typer()
        app.command()(export_example)

        result = runner.invoke(app, [])

        assert result.exit_code == 0
        assert "Available example datasets" in result.stdout

    def test_export_to_yaml(self, tmp_path):
        """Exports example to YAML file."""
        import typer

        app = typer.Typer()
        app.command()(export_example)

        output_path = tmp_path / "out.yaml"
        result = runner.invoke(app, ["miappe/1.1", "-o", str(output_path)])

        assert result.exit_code == 0
        assert output_path.exists()

        # Verify it's valid YAML
        data = yaml.safe_load(output_path.read_text())
        assert data is not None

    def test_export_to_json(self, tmp_path):
        """Exports example to JSON file."""
        import typer

        app = typer.Typer()
        app.command()(export_example)

        output_path = tmp_path / "out.json"
        result = runner.invoke(app, ["miappe/1.1", "-o", str(output_path)])

        assert result.exit_code == 0
        assert output_path.exists()

        # Verify it's valid JSON
        data = json.loads(output_path.read_text())
        assert data is not None

    def test_print_yaml_to_stdout(self, capsys):
        """Prints YAML to stdout when no output specified."""
        import typer

        app = typer.Typer()
        app.command()(export_example)

        result = runner.invoke(app, ["miappe/1.1"])

        assert result.exit_code == 0
        # Should contain YAML content
        assert "unique_id:" in result.stdout or "title:" in result.stdout

    def test_invalid_profile_error(self):
        """Reports error for unknown profile."""
        import typer

        app = typer.Typer()
        app.command()(export_example)

        result = runner.invoke(app, ["nonexistent-profile"])

        assert result.exit_code != 0

    def test_invalid_output_format_error(self, tmp_path):
        """Reports error for unsupported output format."""
        import typer

        app = typer.Typer()
        app.command()(export_example)

        output_path = tmp_path / "out.txt"
        result = runner.invoke(app, ["miappe/1.1", "-o", str(output_path)])

        assert result.exit_code != 0
        # Error message may be in stdout or output
        assert "Unknown output format" in (result.stdout + (result.output or ""))


class TestFlattenEntity:
    """Test entity flattening logic."""

    @pytest.mark.skipif(
        not pytest.importorskip("openpyxl", reason="openpyxl not installed"),
        reason="openpyxl required",
    )
    def test_flattens_simple_fields(self, tmp_path, miappe_example_data):
        """Simple fields are preserved in flattened output."""
        # The flattening is tested indirectly through Excel export
        output_path = tmp_path / "test.xlsx"
        _export_example_to_excel(miappe_example_data, output_path)

        from openpyxl import load_workbook

        wb = load_workbook(output_path)
        inv_sheet = wb["Investigation"]

        # Check header row contains expected fields
        headers = [cell.value for cell in inv_sheet[1]]
        assert "unique_id" in headers
        assert "title" in headers
