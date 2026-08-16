"""A whitespace-only cell must not take the whole Excel export down (260816).

Column width is the longest WORD in the column, because wrapped text folds. A
cell holding only spaces has no words, and `max()` over nothing raises
ValueError — out of `_width`, out of the sheet styling, out of the export. One
stray cell of whitespace, and the user gets no file at all.
"""

from __future__ import annotations

from openpyxl import Workbook

from metaseed.ui.services.sheet_style import _width


def test_a_whitespace_only_cell_is_skipped() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1, value="header")
    sheet.cell(row=2, column=1, value="   ")

    assert _width(sheet, 1, "header", 1) > 0


def test_a_real_value_still_sets_the_width() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1, value="h")
    sheet.cell(row=2, column=1, value="a_long_single_word_value")

    assert _width(sheet, 1, "h", 1) >= len("a_long_single_word_value")
