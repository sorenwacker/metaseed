"""Tests for file parsers."""

import json
from pathlib import Path

import pytest

from metaseed.agent.parsers.csv import CSVParser
from metaseed.agent.parsers.excel import ExcelParser
from metaseed.agent.parsers.json import JSONParser
from metaseed.agent.parsers.registry import create_default_registry


class TestCSVParser:
    """Tests for CSVParser."""

    def test_parse_simple_csv(self, tmp_path: Path) -> None:
        """Parse a simple CSV file."""
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("name,value,active\nfoo,1,true\nbar,2,false\n")

        parser = CSVParser()
        content = parser.parse(csv_file)

        assert content.format == "csv"
        assert len(content.tables) == 1
        assert content.tables[0].headers == ["name", "value", "active"]
        assert content.tables[0].row_count == 2
        assert content.tables[0].rows[0] == ["foo", "1", "true"]

    def test_parse_tsv(self, tmp_path: Path) -> None:
        """Parse a TSV file."""
        tsv_file = tmp_path / "test.tsv"
        tsv_file.write_text("name\tvalue\nfoo\t1\nbar\t2\n")

        parser = CSVParser()
        content = parser.parse(tsv_file)

        assert content.format == "csv"
        assert content.tables[0].headers == ["name", "value"]

    def test_parse_empty_csv(self, tmp_path: Path) -> None:
        """Parse an empty CSV file."""
        csv_file = tmp_path / "empty.csv"
        csv_file.write_text("")

        parser = CSVParser()
        content = parser.parse(csv_file)

        assert len(content.tables) == 0

    def test_can_parse(self, tmp_path: Path) -> None:
        """Check file extension support."""
        parser = CSVParser()

        csv_file = tmp_path / "test.csv"
        csv_file.touch()
        assert parser.can_parse(csv_file) is True

        tsv_file = tmp_path / "test.tsv"
        tsv_file.touch()
        assert parser.can_parse(tsv_file) is True

        json_file = tmp_path / "test.json"
        json_file.touch()
        assert parser.can_parse(json_file) is False


class TestJSONParser:
    """Tests for JSONParser."""

    def test_parse_array(self, tmp_path: Path) -> None:
        """Parse JSON array of objects."""
        json_file = tmp_path / "test.json"
        json_file.write_text(
            json.dumps(
                [
                    {"name": "foo", "value": 1},
                    {"name": "bar", "value": 2},
                ]
            )
        )

        parser = JSONParser()
        content = parser.parse(json_file)

        assert content.format == "json"
        assert len(content.tables) == 1
        assert "name" in content.tables[0].headers
        assert "value" in content.tables[0].headers
        assert content.tables[0].row_count == 2

    def test_parse_object_with_arrays(self, tmp_path: Path) -> None:
        """Parse JSON object containing arrays."""
        json_file = tmp_path / "test.json"
        json_file.write_text(
            json.dumps(
                {
                    "investigations": [
                        {"id": "INV-001", "title": "First"},
                        {"id": "INV-002", "title": "Second"},
                    ],
                    "metadata": "some value",
                }
            )
        )

        parser = JSONParser()
        content = parser.parse(json_file)

        assert content.format == "json"
        assert len(content.tables) == 1
        assert content.tables[0].name == "investigations"
        assert "metadata: some value" in content.text_blocks

    def test_can_parse(self, tmp_path: Path) -> None:
        """Check file extension support."""
        parser = JSONParser()

        json_file = tmp_path / "test.json"
        json_file.touch()
        assert parser.can_parse(json_file) is True

        csv_file = tmp_path / "test.csv"
        csv_file.touch()
        assert parser.can_parse(csv_file) is False


class TestExcelParser:
    """Tests for ExcelParser."""

    def test_can_parse(self, tmp_path: Path) -> None:
        """Check file extension support."""
        parser = ExcelParser()

        xlsx_file = tmp_path / "test.xlsx"
        xlsx_file.touch()
        assert parser.can_parse(xlsx_file) is True

        xls_file = tmp_path / "test.xls"
        xls_file.touch()
        assert parser.can_parse(xls_file) is True

        csv_file = tmp_path / "test.csv"
        csv_file.touch()
        assert parser.can_parse(csv_file) is False

    def test_parse_xlsx(self, tmp_path: Path) -> None:
        """Parse an Excel file."""
        from openpyxl import Workbook

        xlsx_file = tmp_path / "test.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["name", "value", "active"])
        ws.append(["foo", 1, True])
        ws.append(["bar", 2, False])
        wb.save(xlsx_file)

        parser = ExcelParser()
        content = parser.parse(xlsx_file)

        assert content.format == "excel"
        assert len(content.tables) == 1
        assert content.tables[0].name == "Data"
        assert content.tables[0].headers == ["name", "value", "active"]
        assert content.tables[0].row_count == 2

    def test_parse_multiple_sheets(self, tmp_path: Path) -> None:
        """Parse Excel with multiple sheets."""
        from openpyxl import Workbook

        xlsx_file = tmp_path / "multi.xlsx"

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Investigations"
        ws1.append(["id", "title"])
        ws1.append(["INV-001", "First"])

        ws2 = wb.create_sheet("Studies")
        ws2.append(["study_id", "name"])
        ws2.append(["STD-001", "Study One"])

        wb.save(xlsx_file)

        parser = ExcelParser()
        content = parser.parse(xlsx_file)

        assert len(content.tables) == 2
        assert content.tables[0].name == "Investigations"
        assert content.tables[1].name == "Studies"


class TestParserRegistry:
    """Tests for ParserRegistry."""

    def test_create_default_registry(self) -> None:
        """Create registry with default parsers."""
        registry = create_default_registry()

        extensions = registry.supported_extensions()
        assert ".csv" in extensions
        assert ".json" in extensions
        assert ".xlsx" in extensions

    def test_get_parser(self, tmp_path: Path) -> None:
        """Get parser for a file."""
        registry = create_default_registry()

        csv_file = tmp_path / "test.csv"
        csv_file.touch()
        parser = registry.get_parser(csv_file)
        assert isinstance(parser, CSVParser)

        json_file = tmp_path / "test.json"
        json_file.touch()
        parser = registry.get_parser(json_file)
        assert isinstance(parser, JSONParser)

        xlsx_file = tmp_path / "test.xlsx"
        xlsx_file.touch()
        parser = registry.get_parser(xlsx_file)
        assert isinstance(parser, ExcelParser)

    def test_no_parser_found(self, tmp_path: Path) -> None:
        """Raise error when no parser found."""
        registry = create_default_registry()

        unknown_file = tmp_path / "test.unknown"
        unknown_file.touch()

        with pytest.raises(ValueError, match="No parser found"):
            registry.parse(unknown_file)


class TestParsedTable:
    """Tests for ParsedTable."""

    def test_to_dicts(self) -> None:
        """Convert table to list of dicts."""
        from metaseed.agent.parsers.registry import ParsedTable

        table = ParsedTable(
            name="test",
            headers=["a", "b", "c"],
            rows=[
                [1, 2, 3],
                [4, 5, 6],
            ],
        )

        dicts = table.to_dicts()

        assert len(dicts) == 2
        assert dicts[0] == {"a": 1, "b": 2, "c": 3}
        assert dicts[1] == {"a": 4, "b": 5, "c": 6}

    def test_row_count(self) -> None:
        """Get row count."""
        from metaseed.agent.parsers.registry import ParsedTable

        table = ParsedTable(
            name="test",
            headers=["a", "b"],
            rows=[[1, 2], [3, 4], [5, 6]],
        )

        assert table.row_count == 3

    def test_column_count(self) -> None:
        """Get column count."""
        from metaseed.agent.parsers.registry import ParsedTable

        table = ParsedTable(
            name="test",
            headers=["a", "b", "c", "d"],
            rows=[],
        )

        assert table.column_count == 4
