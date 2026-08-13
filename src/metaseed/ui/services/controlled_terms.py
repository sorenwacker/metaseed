"""Controlled terms embedded in an exported workbook, RightField style.

A metadata standard is only followed if following it is the easy path. Exporting
a plain sheet leaves a scientist typing "leaf tissue" where the standard wants
"leaf", filling two hundred rows offline, and finding out on import. Excel can
prevent that at the point of typing: a column restricted to a list offers the
allowed values and refuses anything else, with no network and no knowledge of
ontologies required.

The approach is RightField's (Wolstencroft et al., 2011, *Bioinformatics*
27(14):2021): a curator marks up the template, the scientist fills it in
natively, and the terms' identifiers travel in a hidden sheet so nothing
semantic is lost to a label. Two of its findings are followed here deliberately:

- The identifiers are kept, not just the labels. A label is ambiguous across
  ontologies and unusable as provenance; the IRI is what a downstream system
  can resolve.
- The terms are frozen into the workbook at export. A series of experiments has
  to be annotated against the same vocabulary even if the live ontology moves
  underneath it, so the sheet records the version it captured.

What is *not* attempted: embedding an ontology. A field that names NCBI
Taxonomy cannot carry its two million names into a spreadsheet, and a dropdown
of even a few hundred is worse than useless — the paper's own observation about
shallow hierarchies, arrived at from the other direction. Such a field keeps a
free-text column and the hidden sheet records which ontology the value must
come from, so the requirement is not lost — though nothing validates it yet
(metaseed issue #215): saying otherwise in the sheet would be a promise the
code does not keep.
"""

from __future__ import annotations

from typing import TYPE_CHECKING, Any, NamedTuple

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

    from metaseed.specs.schema import FieldSpec

#: Where the terms and their identifiers live. Hidden, because a scientist
#: filling in the sheet has no use for it, and named so the Inspector-style
#: question "what was this validated against?" has an answer.
TERMS_SHEET = "metaseed terms"

#: The most values worth putting behind a dropdown. Past this, choosing from
#: the list is slower than typing, and the workbook carries weight nobody uses;
#: the field is documented in the terms sheet and checked on import instead.
MAX_EMBEDDED_TERMS = 200

#: How far down a validation reaches. Excel applies validation to a fixed range,
#: so it has to cover rows that do not exist yet — someone adding a row below
#: the exported data must get the same dropdown.
VALIDATED_ROWS = 5000

#: How a cell that breaks the standard looks.
#:
#: The rules warn rather than block: a vocabulary is rarely complete, and
#: someone who knows their value is right should not be locked out of their own
#: spreadsheet by a tool. Excel asks, they may proceed — and the colour stays,
#: so what they accepted is still visible to them and to whoever reads the sheet
#: next. Blocking hides the disagreement; colouring keeps it.
BAD_FILL = PatternFill("solid", fgColor="FBE3E3")
BAD_FONT = Font(color="A32B2B", bold=True)
#: A required value nobody has filled in yet — a reminder, not a mistake.
MISSING_FILL = PatternFill("solid", fgColor="FDF3E0")


class TermList(NamedTuple):
    """The allowed values for one field, with what they came from.

    A list too long to embed is still recorded, with ``embedded`` false: the
    column stays free text, and the sheet says what it must satisfy.

    Attributes:
        entity: Entity type whose column this is.
        field: Field name within that entity.
        values: The allowed values, in the order the specification gives them.
        iris: Identifier per value where the specification knows one, else "".
        source: Where the values came from, for the record in the sheet.
        embedded: Whether the values back a dropdown, or are documentation only.
        note: What a reader of the sheet needs to know, when not embedded.
    """

    entity: str
    field: str
    values: list[str]
    iris: list[str]
    source: str
    embedded: bool = True
    note: str = ""


def allowed_values(field: FieldSpec) -> TermList | None:
    """The controlled values for ``field``, or ``None`` if it has none.

    ``options`` wins over ``constraints.enum`` because it is the marker a
    curator sets deliberately; the enum is the older, validation-side spelling
    of the same idea.
    """
    options = getattr(field, "options", None)
    if options:
        return _term_list(field.name, list(options), "options")

    constraints = getattr(field, "constraints", None)
    enum = getattr(constraints, "enum", None) if constraints else None
    if enum:
        return _term_list(field.name, list(enum), "enum")

    ontologies = getattr(field, "ontologies", None)
    if ontologies:
        # No values to embed — the ontology is not in the workbook and could not
        # be. Say what the column must satisfy, so a reader of the sheet and the
        # import check agree on the requirement.
        return TermList(
            entity="",
            field=field.name,
            values=[],
            iris=[],
            source=", ".join(ontologies),
            embedded=False,
            note=(
                "Free text in the sheet: this vocabulary is too large to embed. "
                "The value should come from the ontology named here; nothing "
                "checks that automatically yet (metaseed issue #215)."
            ),
        )

    return None


def _term_list(field_name: str, values: list[str], source: str) -> TermList:
    """A term list, embedded only if it is short enough to be worth choosing from."""
    if len(values) > MAX_EMBEDDED_TERMS:
        return TermList(
            entity="",
            field=field_name,
            values=values,
            iris=[""] * len(values),
            source=source,
            embedded=False,
            note=(
                f"{len(values)} values: too many for a dropdown, so the column "
                "is free text. The values are listed here, and a wrong one is "
                "caught when the sheet is imported."
            ),
        )
    return TermList("", field_name, values, [""] * len(values), source)


def write_terms_sheet(wb: Workbook, terms: list[TermList]) -> Worksheet | None:
    """Write the hidden sheet holding every term list, and return it.

    Returns ``None`` when nothing is controlled, so a workbook without
    vocabularies gains no empty sheet.
    """
    if not terms:
        return None

    ws = wb.create_sheet(TERMS_SHEET)
    ws.sheet_state = "hidden"

    ws.append(["entity", "field", "value", "identifier", "source", "note"])
    for term_list in terms:
        if not term_list.values:
            ws.append(
                [
                    term_list.entity,
                    term_list.field,
                    "",
                    "",
                    term_list.source,
                    term_list.note,
                ]
            )
            continue
        for value, iri in zip(term_list.values, term_list.iris, strict=False):
            ws.append(
                [
                    term_list.entity,
                    term_list.field,
                    value,
                    iri,
                    term_list.source,
                    term_list.note,
                ]
            )

    for column in range(1, 7):
        ws.column_dimensions[get_column_letter(column)].width = 28
    return ws


def _term_range(ws: Worksheet, term_list: TermList, start_row: int) -> str:
    """The absolute reference to one field's values on the terms sheet."""
    end_row = start_row + len(term_list.values) - 1
    # Quoted because the sheet name contains a space; Excel rejects it otherwise.
    return f"'{ws.title}'!$C${start_row}:$C${end_row}"


def apply_reference_validations(
    facade: Any,
    sheets: dict[str, Worksheet],
    columns_by_entity: dict[str, list[str]],
) -> None:
    """Point every cross-sheet reference at the rows it may name.

    A dataset's sheets are a small relational model: a Study names its
    Investigation, a Sample names the Assay that measured it, and ``_parent``
    names the row above. Typed by hand these are the commonest way an import
    arrives with orphans — a transposed character in an identifier and the row
    attaches to nothing.

    Excel can enforce it: the column is restricted to the identifier column of
    the sheet it points at, so the person filling it in picks an existing row
    instead of retyping its name.
    """
    parents = _parents_by_child(facade)

    for entity, columns in columns_by_entity.items():
        helper = getattr(facade, entity, None)
        ws = sheets.get(entity)
        if helper is None or ws is None:
            continue

        targets: dict[str, tuple[str, str]] = dict(
            getattr(helper, "reference_fields", {}) or {}
        )
        # A reference declared to resolve outside the dataset is deliberately
        # left unrestricted: its value names a GBIF taxon or a museum record,
        # and offering only this dataset's rows would tell the person their
        # correct value is wrong.
        for external in getattr(helper, "external_reference_fields", set()) or set():
            targets.pop(external, None)
        if "_parent" in columns and entity in parents:
            parent_entity, parent_field = parents[entity]
            targets["_parent"] = (parent_entity, parent_field)

        for column, (target_entity, target_field) in targets.items():
            if column not in columns:
                continue
            target_columns = columns_by_entity.get(target_entity)
            target_sheet = sheets.get(target_entity)
            if not target_columns or target_sheet is None:
                continue
            if target_field not in target_columns:
                continue

            source_letter = get_column_letter(columns.index(column) + 1)
            target_letter = get_column_letter(target_columns.index(target_field) + 1)
            validation = DataValidation(
                type="list",
                formula1=(
                    f"'{target_sheet.title}'!"
                    f"${target_letter}$2:${target_letter}${VALIDATED_ROWS}"
                ),
                allow_blank=True,
                showDropDown=False,
                showErrorMessage=True,
                showInputMessage=True,
                errorStyle="warning",
                error=(
                    f"{column} should name a row that exists on the "
                    f"{target_entity} sheet, by its {target_field}. Continue "
                    "anyway if that row is still to come."
                ),
                errorTitle=f"{column}: no such {target_entity}",
                prompt=f"Choose a {target_entity} by its {target_field}.",
                promptTitle=column,
            )
            ws.add_data_validation(validation)
            validation.add(f"{source_letter}2:{source_letter}{VALIDATED_ROWS}")


def flag_invalid_cells(
    facade: Any,
    sheets: dict[str, Worksheet],
    columns_by_entity: dict[str, list[str]],
    fields_by_entity: dict[str, dict[str, Any]],
    terms_by_column: dict[tuple[str, str], str],
) -> None:
    """Mark cells that break the standard, however the value got there.

    Data validation only fires while someone types. Paste a column from another
    workbook — which is how most bulk metadata actually arrives — and Excel
    accepts every value without a murmur. Conditional formatting is evaluated
    continuously instead, so a pasted duplicate or an unknown term is coloured
    the moment it lands.

    Running a check on save would need a macro, which means a file institutions
    block by default; this needs nothing but the sheet.
    """
    for entity, columns in columns_by_entity.items():
        ws = sheets.get(entity)
        helper = getattr(facade, entity, None)
        if ws is None or helper is None:
            continue

        fields = fields_by_entity.get(entity, {})
        keys = key_columns(facade, entity, fields)

        for column in columns:
            letter = get_column_letter(columns.index(column) + 1)
            cells = f"{letter}2:{letter}{VALIDATED_ROWS}"
            field = fields.get(column)

            if column in keys:
                ws.conditional_formatting.add(
                    cells,
                    FormulaRule(
                        formula=[
                            f'AND({letter}2<>"",COUNTIF(${letter}$2:${letter}${VALIDATED_ROWS},{letter}2)>1)'
                        ],
                        stopIfTrue=False,
                        fill=BAD_FILL,
                        font=BAD_FONT,
                    ),
                )

            term_range = terms_by_column.get((entity, column))
            if term_range:
                ws.conditional_formatting.add(
                    cells,
                    FormulaRule(
                        formula=[
                            f'AND({letter}2<>"",COUNTIF({term_range},{letter}2)=0)'
                        ],
                        stopIfTrue=False,
                        fill=BAD_FILL,
                        font=BAD_FONT,
                    ),
                )

            if getattr(field, "required", False):
                # Empty only counts against a row someone has started: a blank
                # row below the data is not an error, it is unfilled.
                first = get_column_letter(1)
                ws.conditional_formatting.add(
                    cells,
                    FormulaRule(
                        formula=[f'AND(${first}2<>"",{letter}2="")'],
                        stopIfTrue=False,
                        fill=MISSING_FILL,
                    ),
                )


def key_columns(facade: Any, entity: str, fields: dict[str, Any]) -> set[str]:
    """Columns of ``entity`` that genuinely have to be unique.

    Three things can make a column a key, and the facade's ``identifier_field``
    is not one of them on its own: it falls back to the first field when a
    profile declares nothing, which is how ``File.filename`` and attribute-style
    ``tag`` columns came to be treated as identifiers and flagged on every row
    that legitimately repeated.

    A column is a key when the specification says so — ``unique_within`` or the
    identifier marker — or when something points at it, because a reference
    that resolves to two rows resolves to neither.
    """
    keys = {
        name
        for name, field in fields.items()
        if getattr(field, "unique_within", None)
        or getattr(field, "is_identifier", False)
    }

    for other in getattr(facade, "entities", []) or []:
        helper = getattr(facade, other, None)
        for target_entity, target_field in (
            getattr(helper, "reference_fields", {}) or {}
        ).values():
            if target_entity == entity:
                keys.add(target_field)

    return keys


def apply_uniqueness_validations(
    facade: Any,
    sheets: dict[str, Worksheet],
    columns_by_entity: dict[str, list[str]],
    fields_by_entity: dict[str, dict[str, Any]],
) -> None:
    """Refuse a duplicate identifier where the specification requires one.

    An identifier that repeats is not a typo the import can resolve: two rows
    claim the same name, every reference to it becomes ambiguous, and the
    second row silently replaces the first. Excel can catch it as it is typed —
    a COUNTIF over the column, refusing anything already present.

    This does not replace the check on import. Excel applies validation to
    typing, not to pasting, and a pasted block is exactly how a column of
    duplicates arrives.
    """
    for entity, columns in columns_by_entity.items():
        ws = sheets.get(entity)
        helper = getattr(facade, entity, None)
        if ws is None or helper is None:
            continue

        fields = fields_by_entity.get(entity, {})
        unique_columns = {
            name
            for name, field in fields.items()
            if getattr(field, "unique_within", None)
            or getattr(field, "is_identifier", False)
        }
        identifier = getattr(helper, "identifier_field", None)
        if identifier:
            unique_columns.add(identifier)

        for column in unique_columns & set(columns):
            letter = get_column_letter(columns.index(column) + 1)
            validation = DataValidation(
                type="custom",
                formula1=f"COUNTIF(${letter}$2:${letter}${VALIDATED_ROWS},{letter}2)<2",
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="warning",
                error=(
                    f"{column} is already used on another row. Each row needs "
                    "its own, or references to it cannot say which row they "
                    "mean. Continue anyway if you know better — the cell stays "
                    "marked."
                ),
                errorTitle=f"{column}: already used",
            )
            ws.add_data_validation(validation)
            validation.add(f"{letter}2:{letter}{VALIDATED_ROWS}")


def _parents_by_child(facade: Any) -> dict[str, tuple[str, str]]:
    """Child entity -> the entity that nests it and that entity's identifier.

    Read from the nesting the profile declares, so ``_parent`` points at
    whatever the specification says holds this entity rather than a guess.
    """
    parents: dict[str, tuple[str, str]] = {}
    for parent in getattr(facade, "entities", []) or []:
        parent_helper = getattr(facade, parent, None)
        if parent_helper is None:
            continue
        identifier = getattr(parent_helper, "identifier_field", None)
        if not identifier:
            continue
        for child in (getattr(parent_helper, "nested_fields", {}) or {}).values():
            parents.setdefault(child, (parent, identifier))
    return parents


def apply_validations(
    wb: Workbook,
    sheets: dict[str, Worksheet],
    columns_by_entity: dict[str, list[str]],
    fields_by_entity: dict[str, dict[str, Any]],
    term_ranges: dict[tuple[str, str], str] | None = None,
) -> Worksheet | None:
    """Restrict every controlled column to its terms, and record them.

    Args:
        wb: The workbook being built.
        sheets: Entity type -> its worksheet.
        columns_by_entity: Entity type -> the columns written, in order.
        fields_by_entity: Entity type -> field name -> its :class:`FieldSpec`.

    Returns:
        The terms sheet, or ``None`` when no field in the workbook is
        controlled. ``term_ranges`` is filled in with where each embedded list
        lives, so the same values can drive the conditional formatting.
    """
    term_ranges = {} if term_ranges is None else term_ranges
    collected: list[tuple[TermList, str, int]] = []
    row = 2  # after the header
    for entity, columns in columns_by_entity.items():
        fields = fields_by_entity.get(entity, {})
        for column in columns:
            field = fields.get(column)
            if field is None:
                continue
            terms = allowed_values(field)
            if terms is None:
                continue
            terms = terms._replace(entity=entity)
            collected.append((terms, column, row))
            row += max(len(terms.values), 1)

    if not collected:
        return None

    terms_sheet = write_terms_sheet(wb, [t for t, _, _ in collected])
    if terms_sheet is None:  # pragma: no cover - collected is non-empty here
        return None

    for terms, column, start_row in collected:
        if not terms.embedded:
            continue  # documented in the sheet; the column stays free text
        term_ranges[(terms.entity, column)] = _term_range(terms_sheet, terms, start_row)
        ws = sheets[terms.entity]
        letter = get_column_letter(columns_by_entity[terms.entity].index(column) + 1)
        validation = DataValidation(
            type="list",
            formula1=_term_range(terms_sheet, terms, start_row),
            allow_blank=True,
            showDropDown=False,  # False *shows* the arrow; True hides it (ECMA-376)
            showErrorMessage=True,  # without this Excel accepts anything typed
            showInputMessage=True,
            errorStyle="warning",
            error=(
                f"{column} takes one of the values listed for it. Pick from the "
                "dropdown, or continue with your own value — the cell stays "
                "marked so the difference is visible."
            ),
            errorTitle=f"{column}: value not in the standard",
            prompt="Choose a value from the list.",
            promptTitle=column,
        )
        ws.add_data_validation(validation)
        validation.add(f"{letter}2:{letter}{VALIDATED_ROWS}")

    return terms_sheet
