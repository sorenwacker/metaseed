"""Excel export service.

Builds Excel workbook from AppState entity tree.
"""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook

from metaseed.ui.helpers import to_dict
from metaseed.ui.services.controlled_terms import (
    apply_reference_validations,
    apply_uniqueness_validations,
    apply_validations,
    flag_invalid_cells,
)
from metaseed.ui.services.sheet_style import style_sheet

if TYPE_CHECKING:
    from metaseed.ui.state import AppState

# Characters that make Excel/LibreOffice interpret a cell as a formula. A string
# value beginning with one of these (e.g. a collaborator-supplied field like
# ``=HYPERLINK(...)``) would otherwise round-trip into a live formula on export.
FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _escape_formula(value: object) -> object:
    """Neutralize a formula-injection payload in a string cell value.

    Prefixes a single quote so the value is stored/opened as literal text, not a
    formula (also stops openpyxl from emitting a leading-``=`` string as a
    formula). Non-strings can't be formulas and are returned unchanged.
    """
    if isinstance(value, str) and value.startswith(FORMULA_TRIGGERS):
        return "'" + value
    return value


def _format_cell_value(value: object, is_nested_field: bool) -> object:
    """Format a value for Excel cell.

    Args:
        value: The value to format.
        is_nested_field: Whether this field contains nested entities.

    Returns:
        Formatted value suitable for Excel.
    """
    if is_nested_field:
        if isinstance(value, list):
            return len(value)
        if value:
            return 1
        return 0
    if isinstance(value, list):
        if value and not isinstance(value[0], dict):
            return ", ".join(str(v) for v in value)
        # An empty scalar list must export as an empty cell: "0" would fail
        # list validation on reimport and silently drop the whole entity.
        return len(value) if value else ""
    if isinstance(value, dict):
        return "[object]"
    if not isinstance(value, str | int | float | bool | type(None)):
        return str(value)
    return value


def _stated_values(data: dict[str, Any]) -> dict[str, str]:
    """The scalar values a row actually states, as strings.

    Nested lists and blanks are left out: a stored row carries every field the
    model declares, most of them empty, while an embedded copy carries only
    what was written.
    """
    return {
        key: str(value)
        for key, value in data.items()
        if key != "_parent"
        and not isinstance(value, list | dict)
        and value not in (None, "")
    }


def _is_contained_in(candidate: dict[str, Any], rows: list[dict[str, Any]]) -> bool:
    """Whether ``candidate`` states nothing that some row does not already say.

    Containment rather than equality, because the stored row is the fuller of
    the two. Identifiers cannot decide this: entities repeat one legitimately —
    every sample carries an attribute tagged ``collection date`` — and matching
    on that deletes real rows.
    """
    stated = _stated_values(candidate)
    return any(
        all(_stated_values(row).get(key) == value for key, value in stated.items())
        for row in rows
    )


def _stored_rows(facade: Any) -> list[tuple[str, dict[str, Any]]]:
    """Every stored entity as ``(type, data)``, carrying its parent."""
    from metaseed import MetaseedClient

    rows: list[tuple[str, dict[str, Any]]] = []
    payload = MetaseedClient.from_facade(facade).serialize()
    for entity in payload.get("entities", []):
        entity_type = entity.get("_type")
        if not entity_type:
            continue
        data = {k: v for k, v in entity.items() if not k.startswith("_")}
        if entity.get("_parent_unique_id"):
            # The tree, as a business key. No profile declares parent_ref
            # fields, so without this column the export cannot be reimported
            # with its structure intact.
            data["_parent"] = entity["_parent_unique_id"]
        rows.append((entity_type, data))
    return rows


def _identifier_of(facade: Any, entity_type: str, data: dict[str, Any]) -> str:
    helper = getattr(facade, entity_type, None)
    field = getattr(helper, "identifier_field", None) if helper else None
    return str(data.get(field, "")) if field else ""


def collect_rows_by_type(facade: Any) -> dict[str, list[dict[str, Any]]]:
    # Renamed from collect_entities_by_type: helpers/entity_helpers.py exports
    # a same-named function with a different signature and meaning (dropdown
    # entries), and two public names that differ only by import path invite
    # the wrong one.
    """Group every entity in ``facade`` by its type, each appearing once.

    An entity can be present twice over: as a stored row of its own, and as the
    dict still embedded in its parent's data. Emitting both put every child in
    the sheet twice — with the copy carrying no ``_parent``, since only a
    stored row knows what it hangs from — which read as a column of duplicate
    identifiers and a row belonging to nothing.

    Stored rows are taken first and an embedded dict is emitted only when no
    stored row under the same parent already contains it, which is what happens
    for a child that was never materialised.
    """
    by_type: dict[str, list[dict[str, Any]]] = {}
    emitted: dict[tuple[str, str], list[dict[str, Any]]] = {}

    def emit(entity_type: str, parent: str, data: dict[str, Any]) -> None:
        by_type.setdefault(entity_type, []).append(data)
        emitted.setdefault((entity_type, parent), []).append(data)

    def walk_embedded(entity_type: str, data: dict[str, Any]) -> None:
        helper = getattr(facade, entity_type, None)
        if helper is None:
            return
        parent = _identifier_of(facade, entity_type, data)
        for field_name, nested_type in helper.nested_fields.items():
            for item in data.get(field_name) or []:
                if not isinstance(item, dict):
                    continue
                if not _is_contained_in(item, emitted.get((nested_type, parent), [])):
                    emit(nested_type, parent, item)
                walk_embedded(nested_type, item)

    rows = _stored_rows(facade)
    for entity_type, data in rows:
        emit(entity_type, str(data.get("_parent", "")), data)
    for entity_type, data in rows:
        walk_embedded(entity_type, data)

    return by_type


def build_workbook_from_facade(facade: Any) -> Workbook:
    """Build the workbook for a dataset, from the facade holding it.

    Takes the facade rather than an application's own state object, because
    that is the one thing every caller has: the hub kept a copy of this
    function for exactly that reason, and the copy did not gain the dropdowns,
    the tables or the descriptions this one has.
    """
    entities_by_type = collect_rows_by_type(facade)

    # Children live on their own sheet, linked by ``_parent``; a parent's own
    # copy of its nested list is emptied by the flat serialization. So count the
    # children that name each parent, keyed by (child type, parent identifier) --
    # reading the emptied list instead reported 0 for every parent.
    child_counts: dict[tuple[str, str], int] = {}
    for child_type, child_rows in entities_by_type.items():
        for child_row in child_rows:
            parent_id = child_row.get("_parent")
            if parent_id:
                key = (child_type, str(parent_id))
                child_counts[key] = child_counts.get(key, 0) + 1

    wb = Workbook()
    wb.remove(wb.active)

    # Kept so the controlled columns can be restricted once every sheet exists.
    sheets: dict[str, Any] = {}
    columns_by_entity: dict[str, list[str]] = {}
    fields_by_entity: dict[str, dict[str, Any]] = {}

    for entity_type in facade.entities:
        helper = getattr(facade, entity_type, None)
        if not helper:
            continue

        ws = wb.create_sheet(entity_type)
        sheets[entity_type] = ws
        nested_types = helper.nested_fields  # field name -> contained entity type
        columns = [*helper.all_fields, "_parent"]
        columns_by_entity[entity_type] = columns
        fields_by_entity[entity_type] = _field_specs(facade, entity_type)

        ws.append(columns)

        entities = entities_by_type.get(entity_type, [])
        for row_offset, entity_data in enumerate(entities, start=2):
            parent_id = _identifier_of(facade, entity_type, entity_data)
            for col_offset, col in enumerate(columns, start=1):
                if col in nested_types:
                    # How many children of this type hang from this row.
                    value: object = child_counts.get((nested_types[col], parent_id), 0)
                else:
                    value = _format_cell_value(entity_data.get(col, ""), False)
                value = _escape_formula(value)
                cell = ws.cell(
                    row=row_offset,
                    column=col_offset,
                    value=str(value) if value != "" else "",
                )
                # Every data cell is text. Excel otherwise reinterprets what it
                # recognises -- gene names become dates, identifiers lose their
                # leading zeros -- and a metadata value must survive the round
                # trip byte for byte.
                cell.number_format = "@"

        style_sheet(ws, columns, fields_by_entity[entity_type], len(entities))

    # A column the specification controls becomes a dropdown, and the terms are
    # written into a hidden sheet with what they came from. See
    # metaseed.ui.services.controlled_terms.
    term_ranges: dict[tuple[str, str], str] = {}
    apply_validations(wb, sheets, columns_by_entity, fields_by_entity, term_ranges)
    apply_reference_validations(facade, sheets, columns_by_entity)
    apply_uniqueness_validations(facade, sheets, columns_by_entity, fields_by_entity)
    # Validation stops at the keyboard; pasted values need a check that keeps
    # looking.
    flag_invalid_cells(facade, sheets, columns_by_entity, fields_by_entity, term_ranges)

    return wb


def build_workbook(state: AppState) -> Workbook:
    """Build the workbook for the dataset in ``state``."""
    return build_workbook_from_facade(state.get_or_create_facade())


def _field_specs(facade: Any, entity_type: str) -> dict[str, Any]:
    """Field name -> its :class:`FieldSpec` for one entity, or empty if unknown.

    A workbook must still export when the specification cannot be loaded, just
    without the dropdowns and descriptions that come from it.
    """
    from metaseed.specs.loader import SpecLoader

    try:
        spec = SpecLoader().load_profile(facade.version, facade.profile)
    except Exception:
        return {}
    entity = spec.entities.get(entity_type)
    fields = getattr(entity, "fields", None)
    return {field.name: field for field in fields} if fields else {}


def export_to_bytes(state: AppState) -> BytesIO:
    """Export entity tree to Excel bytes.

    Args:
        state: The current AppState containing the entity tree.

    Returns:
        BytesIO containing the Excel file.
    """
    wb = build_workbook(state)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_filename(state: AppState) -> str:
    """Generate filename for the export.

    Args:
        state: The current AppState containing the entity tree.

    Returns:
        Filename string for the Excel export.
    """
    facade = state.get_or_create_facade()

    date_str = datetime.now(UTC).strftime("%y%m%d")
    version_str = facade.version.replace(".", "-")

    entity_id = "export"
    root_nodes = [n for n in state.nodes_by_id.values() if n.parent_id is None]
    if root_nodes:
        root_node = root_nodes[0]
        root_data = to_dict(root_node.instance) or {}
        if root_data.get("unique_id"):
            entity_id = (
                str(root_data["unique_id"]).replace("/", "-").replace(":", "-")[:30]
            )

    return f"{date_str}-{state.profile}-{version_str}-{entity_id}.xlsx"
