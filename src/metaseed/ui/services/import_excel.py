"""Excel import: the reverse of the Excel export.

Reads a workbook shaped the way :mod:`metaseed.ui.services.export` writes one —
one sheet per entity type, a header row of field names, every cell text — back
into the dataset payload :func:`metaseed.ui.datasets.import_payload` loads, so
Excel and JSON imports share one loading path.

The tree comes from the ``_parent`` column the export writes: no profile
declares ``parent_ref`` fields, so without that column the linkage would not
survive the round trip.

Nested list fields (a parent's embedded children) are skipped on import: the
children arrive as rows on their own sheet, and the export flattens the list to
a count that means nothing on the way back.
"""

from __future__ import annotations

from io import BytesIO
from typing import TYPE_CHECKING, Any

from openpyxl import load_workbook

from metaseed.specs.loader import SpecLoader
from metaseed.specs.schema import FieldType
from metaseed.ui.services.export import FORMULA_TRIGGERS

if TYPE_CHECKING:
    from metaseed.facade import ProfileFacade

#: The column carrying the parent's identifier, written by the export.
PARENT_COLUMN = "_parent"


def _unescape_formula(value: str) -> str:
    """Reverse the export's formula neutralisation.

    The export prefixes formula-triggering cells with a quote so Excel shows
    them as text; importing must give back the original value or a round trip
    changes the data.
    """
    if (
        value.startswith("'")
        and len(value) > 1
        and value[1:].startswith(FORMULA_TRIGGERS)
    ):
        return value[1:]
    return value


def workbook_to_payload(
    raw: bytes, *, profile: str, version: str, facade: ProfileFacade
) -> dict[str, Any]:
    """Parse an exported workbook back into a loadable dataset payload.

    Args:
        raw: The ``.xlsx`` file contents.
        profile: Profile the dataset belongs to (the workbook does not carry it).
        version: Profile version, for the same reason.
        facade: The profile facade, used to know each sheet's entity type and
            which columns are nested lists to skip.

    Returns:
        A payload for :func:`metaseed.ui.datasets.import_payload`.

    Raises:
        ValueError: If the file is not a readable workbook, or no sheet matches
            an entity type of the profile — which means it was exported from a
            different profile, and loading it would silently produce nothing.
    """
    try:
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Not a readable Excel workbook: {exc}") from exc

    # The spec says which columns are entity lists (skip: children arrive on
    # their own sheets) and which are scalar lists (split what export joined).
    spec = SpecLoader().load_profile(version, profile)

    entities: list[dict[str, Any]] = []
    # The export's own sheets are not entities; matching by entity name already
    # skips them, and naming the fact keeps it true if that changes.
    for entity_type in facade.entities:
        if entity_type not in workbook.sheetnames:
            continue
        fields = (
            {f.name: f for f in spec.entities[entity_type].fields}
            if entity_type in spec.entities
            else {}
        )
        nested = {n for n, f in fields.items() if f.is_nested()}
        scalar_lists = {
            n for n, f in fields.items() if f.type == FieldType.LIST and n not in nested
        }

        rows = workbook[entity_type].iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            continue
        columns = [str(c) if c is not None else "" for c in header]

        for row in rows:
            data: dict[str, Any] = {"_type": entity_type}
            for column, cell in zip(columns, row, strict=False):
                if cell is None or cell == "" or not column:
                    continue
                value = _unescape_formula(str(cell))
                if column == PARENT_COLUMN:
                    data["_parent_unique_id"] = value
                elif column in scalar_lists:
                    data[column] = [v.strip() for v in value.split(",") if v.strip()]
                elif column not in nested:
                    data[column] = value
            if len(data) > 1:  # a row of empties is not an entity
                entities.append(data)

    if not entities:
        # An empty result is never what an import meant. The usual cause is a
        # workbook exported from a different profile: sheet names largely miss,
        # and any that coincide hold no rows this profile can read.
        raise ValueError(
            f"Nothing in this workbook matches the {profile} profile — was it "
            "exported from a different profile?"
        )

    return {"profile": profile, "version": version, "entities": entities}
