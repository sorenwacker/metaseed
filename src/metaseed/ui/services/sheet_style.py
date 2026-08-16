"""Making an exported sheet readable to the person who has to fill it in.

A metadata sheet is not a data dump: someone reads every column heading and
decides what belongs under it. Exported plain, the headings are truncated
machine names in a grid that scrolls away from them, and what each field means
lives in a specification the person filling the sheet has never opened.

So: the heading row and the identifier column stay on screen, headings are
legible and marked by whether they are required, each carries its description
from the specification as a comment, columns are wide enough to read, long
values wrap instead of running under the next cell, and alternate rows are
tinted so the eye holds its line across a wide sheet.

Nothing here changes a value. The import reads row 1 as the heading and every
row below as data, which is why descriptions are comments rather than a second
row — a sheet that documents itself must not become a sheet that imports its own
documentation.
"""

from __future__ import annotations

from typing import TYPE_CHECKING, Any

from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

#: The palette the application uses, so an exported sheet is recognisably from
#: the same tool rather than Excel's defaults.
HEADER_FILL = PatternFill("solid", fgColor="1A3A2F")
REQUIRED_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
OPTIONAL_FONT = Font(name="Calibri", size=11, bold=False, color="E8EFE9")
HEADER_HEIGHT = 34

#: Column width bounds. Narrow enough that a sheet of them still fits a screen,
#: wide enough that a heading is not cut off mid-word.
MIN_WIDTH = 12
MAX_WIDTH = 42

#: Each entity sheet is a real Excel table, not a range that has been painted
#: to look like one. A table bands its rows itself, and — the reason it is worth
#: it — a row typed directly beneath one is absorbed into it, inheriting the
#: banding *and* the column's data validation. Painted banding stops where the
#: data stopped, and the dropdown on an added row would have to be guessed at by
#: covering thousands of empty rows in advance.
TABLE_STYLE = TableStyleInfo(
    name="TableStyleLight21",  # the green family, closest to the app's palette
    showRowStripes=True,
    showColumnStripes=False,
    showFirstColumn=False,
    showLastColumn=False,
)

#: The structural column is not metadata and nobody should be inventing values
#: for it: it is written by the export and read back by the import. It stays
#: visible because a row added in Excel has no other way to say which parent it
#: belongs to — and its dropdown lists exactly the parents that exist.
SYSTEM_FILL = PatternFill("solid", fgColor="ECEFE9")
SYSTEM_HEADER_FONT = Font(name="Calibri", size=11, italic=True, color="C9D3C6")
SYSTEM_FONT = Font(name="Calibri", size=11, italic=True, color="6B7A70")
SYSTEM_COLUMNS = frozenset({"_parent"})

#: One empty row is kept inside the table when an entity has no data: a table
#: cannot be a header alone, and it gives the person a row to type into that is
#: already part of the table.
EMPTY_TABLE_ROWS = 1


def style_sheet(
    ws: Worksheet,
    columns: list[str],
    fields: dict[str, Any],
    row_count: int,
) -> None:
    """Lay out one entity sheet: headings, widths, wrapping, frozen panes.

    Args:
        ws: The worksheet holding this entity's rows.
        columns: Column names in the order they were written.
        fields: Column name -> its :class:`FieldSpec`, where the specification
            knows the column. ``_parent`` and unknown columns simply get no
            description.
        row_count: How many data rows were written, for sizing the columns to
            what is actually in them.
    """
    for index, column in enumerate(columns, start=1):
        letter = get_column_letter(index)
        field = fields.get(column)
        cell = ws.cell(row=1, column=index)

        required = bool(getattr(field, "required", False))
        if column in SYSTEM_COLUMNS:
            cell.font = SYSTEM_HEADER_FONT
        else:
            cell.font = REQUIRED_FONT if required else OPTIONAL_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        note = _description(column, field)
        if note:
            # A comment, not a second row: the import reads everything below the
            # heading as data.
            cell.comment = Comment(note, "metaseed", width=320, height=160)

        ws.column_dimensions[letter].width = _width(ws, index, column, row_count)

    ws.row_dimensions[1].height = HEADER_HEIGHT

    # Row 1 and the first column stay put: the first column carries the
    # identifier, without which a row scrolled sideways cannot be placed.
    ws.freeze_panes = "B2"

    system_columns = {
        index
        for index, column in enumerate(columns, start=1)
        if column in SYSTEM_COLUMNS
    }

    last_row = max(row_count + 1, 1 + EMPTY_TABLE_ROWS)
    for row in ws.iter_rows(min_row=2, max_row=last_row, max_col=len(columns)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.column in system_columns:
                # Set apart from the metadata: this is structure, not something
                # observed or measured. A fill here beats the table's banding,
                # which is the intent.
                cell.fill = SYSTEM_FILL
                cell.font = SYSTEM_FONT

    _make_table(ws, columns, last_row)


def _description(column: str, field: Any) -> str:
    """What to tell someone hovering over this heading.

    The specification's own description, plus the facts that decide what may go
    in the cell: whether it is required, its unit, and an example.
    """
    if column == "_parent":
        return (
            "Structure, not metadata: which row of the parent sheet this row "
            "belongs to, named by that row's identifier.\n\n"
            "Written by the export — leave the existing values alone. If you "
            "add a row, choose its parent from the dropdown, or the row will "
            "have nothing to attach to on import."
        )
    if field is None:
        return ""

    parts: list[str] = []
    description = (getattr(field, "description", "") or "").strip()
    if description:
        parts.append(description)

    facts: list[str] = []
    facts.append("Required" if getattr(field, "required", False) else "Optional")
    unit = getattr(field, "unit", None)
    if unit:
        facts.append(f"Unit: {unit}")
    example = getattr(field, "example", None)
    if example:
        facts.append(f"Example: {example}")
    parts.append(" · ".join(facts))

    return "\n\n".join(parts)


def _width(ws: Worksheet, index: int, column: str, row_count: int) -> float:
    """A column wide enough for its heading and its values, within reason.

    Sized from what the sheet actually holds rather than a fixed guess: a column
    of dates and a column of abstracts do not want the same width, and a
    heading that is cut off is a question nobody can answer.
    """
    widest = len(column)
    for row in range(2, min(row_count, 200) + 2):
        value = ws.cell(row=row, column=index).value
        if value:
            # Wrapped text needs the width of its longest word, not its whole
            # length: the rest folds onto the next line. A whitespace-only cell
            # has no words, and max() of nothing took the whole export down.
            words = str(value).split()
            if not words:
                continue
            widest = max(widest, min(len(str(value)), max(len(w) for w in words)))
    return max(MIN_WIDTH, min(MAX_WIDTH, widest + 3))


def _make_table(ws: Worksheet, columns: list[str], last_row: int) -> None:
    """Turn the written range into an Excel table.

    The name has to be unique in the workbook and cannot contain spaces, so it
    is derived from the sheet rather than the entity's display name.
    """
    last_letter = get_column_letter(len(columns))
    table = Table(
        displayName=f"tbl_{ws.title.replace(' ', '_')}",
        ref=f"A1:{last_letter}{last_row}",
    )
    table.tableStyleInfo = TABLE_STYLE
    ws.add_table(table)
