"""CLI command for exporting example data."""

from pathlib import Path
from typing import Annotated, Any

import typer
import yaml

from metaseed.cli.output import echo_error, echo_success

# Exit codes
EXIT_SUCCESS = 0
EXIT_VALIDATION_ERROR = 1
EXIT_INPUT_ERROR = 2
EXIT_CONFIG_ERROR = 3


def _export_example_to_excel(data: dict[str, Any], output: Path) -> None:
    """Export example data to Excel with multiple sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    def flatten_entity(
        entity_data: dict[str, Any], parent_fields: dict[str, Any] | None = None
    ) -> dict[str, Any]:
        """Flatten nested entity, excluding nested lists."""
        flat: dict[str, Any] = {}
        if parent_fields:
            flat.update(parent_fields)
        for key, value in entity_data.items():
            if not isinstance(value, list | dict) or key in [
                "associated_publications",
                "external_references",
            ]:
                flat[key] = (
                    value if not isinstance(value, list) else ", ".join(str(v) for v in value)
                )
        return flat

    def add_sheet(sheet_name: str, records: list[dict[str, Any]]) -> None:
        """Add a sheet with records."""
        if not records:
            return

        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit

        # Collect all unique keys preserving order
        all_keys: list[str] = []
        for record in records:
            for key in record:
                if key not in all_keys:
                    all_keys.append(key)

        # Write header
        for col, key in enumerate(all_keys, 1):
            cell = ws.cell(row=1, column=col, value=key)
            cell.font = header_font
            cell.fill = header_fill

        # Write data
        for row, record in enumerate(records, 2):
            for col, key in enumerate(all_keys, 1):
                value = record.get(key, "")
                if isinstance(value, list | dict):
                    value = str(value)
                ws.cell(row=row, column=col, value=value)

        # Auto-width columns
        for col in range(1, len(all_keys) + 1):
            max_length = max(
                len(str(ws.cell(row=row, column=col).value or ""))
                for row in range(1, len(records) + 2)
            )
            ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)

    # Root entity (Investigation)
    root_record = flatten_entity(data)
    add_sheet("Investigation", [root_record])

    # Contacts/Persons at investigation level
    if data.get("contacts"):
        records = [flatten_entity(c) for c in data["contacts"]]
        add_sheet("Person", records)

    # Studies
    if data.get("studies"):
        study_records = []
        all_persons: list[dict[str, Any]] = []
        all_locations: list[dict[str, Any]] = []
        all_bio_materials: list[dict[str, Any]] = []
        all_factors: list[dict[str, Any]] = []
        all_factor_values: list[dict[str, Any]] = []
        all_obs_variables: list[dict[str, Any]] = []
        all_obs_units: list[dict[str, Any]] = []
        all_samples: list[dict[str, Any]] = []
        all_events: list[dict[str, Any]] = []
        all_environments: list[dict[str, Any]] = []
        all_data_files: list[dict[str, Any]] = []
        all_protocols: list[dict[str, Any]] = []
        all_sources: list[dict[str, Any]] = []
        all_assays: list[dict[str, Any]] = []

        for study in data["studies"]:
            study_records.append(flatten_entity(study))

            # Nested entities within study
            if "persons" in study:
                all_persons.extend(flatten_entity(p) for p in study["persons"])
            if study.get("geographic_location"):
                loc = study["geographic_location"]
                if isinstance(loc, dict):
                    all_locations.append(flatten_entity(loc))
            if "biological_materials" in study:
                for bm in study["biological_materials"]:
                    all_bio_materials.append(flatten_entity(bm))
            if "factors" in study:
                for f in study["factors"]:
                    all_factors.append(flatten_entity(f))
                    if "values" in f:
                        all_factor_values.extend(flatten_entity(fv) for fv in f["values"])
            if "observed_variables" in study:
                all_obs_variables.extend(flatten_entity(ov) for ov in study["observed_variables"])
            if "observation_units" in study:
                for ou in study["observation_units"]:
                    all_obs_units.append(flatten_entity(ou))
                    if "samples" in ou:
                        all_samples.extend(flatten_entity(s) for s in ou["samples"])
                    if "factor_values" in ou:
                        all_factor_values.extend(flatten_entity(fv) for fv in ou["factor_values"])
            if "events" in study:
                all_events.extend(flatten_entity(e) for e in study["events"])
            if "environments" in study:
                all_environments.extend(flatten_entity(env) for env in study["environments"])
            if "data_files" in study:
                all_data_files.extend(flatten_entity(df) for df in study["data_files"])
            if "protocols" in study:
                all_protocols.extend(flatten_entity(p) for p in study["protocols"])
            if "sources" in study:
                all_sources.extend(flatten_entity(s) for s in study["sources"])
            if "samples" in study:
                all_samples.extend(flatten_entity(s) for s in study["samples"])
            if "assays" in study:
                all_assays.extend(flatten_entity(a) for a in study["assays"])

        add_sheet("Study", study_records)
        if all_persons:
            add_sheet("Person", all_persons)
        if all_locations:
            add_sheet("Location", all_locations)
        if all_bio_materials:
            add_sheet("BiologicalMaterial", all_bio_materials)
        if all_factors:
            add_sheet("Factor", all_factors)
        if all_factor_values:
            add_sheet("FactorValue", all_factor_values)
        if all_obs_variables:
            add_sheet("ObservedVariable", all_obs_variables)
        if all_obs_units:
            add_sheet("ObservationUnit", all_obs_units)
        if all_samples:
            add_sheet("Sample", all_samples)
        if all_events:
            add_sheet("Event", all_events)
        if all_environments:
            add_sheet("Environment", all_environments)
        if all_data_files:
            add_sheet("DataFile", all_data_files)
        if all_protocols:
            add_sheet("Protocol", all_protocols)
        if all_sources:
            add_sheet("Source", all_sources)
        if all_assays:
            add_sheet("Assay", all_assays)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def export_example(
    profile: Annotated[
        str | None, typer.Argument(help="Profile name (miappe, isa, isa-miappe-combined)")
    ] = None,
    output: Annotated[
        Path | None, typer.Option("--output", "-o", help="Output file path (.xlsx or .yaml)")
    ] = None,
    list_examples: Annotated[
        bool, typer.Option("--list", "-l", help="List available examples")
    ] = False,
) -> None:
    """Export example data for a profile.

    Examples are fully populated sample datasets demonstrating each profile's
    entity types and relationships.

    Usage:
        metaseed example --list              # List available examples
        metaseed example miappe -o out.xlsx  # Export MIAPPE example to Excel
        metaseed example isa -o out.yaml     # Export ISA example to YAML
    """
    import importlib.resources
    import json

    # Find examples directory (in src/metaseed/examples/)
    try:
        examples_dir = Path(importlib.resources.files("metaseed")) / "examples"
        if not examples_dir.exists():
            # Try relative to package (from cli/commands/example.py -> metaseed/examples)
            examples_dir = Path(__file__).parent.parent.parent / "examples"
    except Exception:
        examples_dir = Path(__file__).parent.parent.parent / "examples"

    if not examples_dir.exists():
        echo_error(f"Examples directory not found at {examples_dir}")
        raise typer.Exit(EXIT_CONFIG_ERROR)

    # Get available examples (profile/version/example.yaml structure)
    example_files: dict[str, Path] = {}
    if examples_dir.exists():
        for profile_dir in examples_dir.iterdir():
            if profile_dir.is_dir() and not profile_dir.name.startswith("."):
                for version_dir in profile_dir.iterdir():
                    if version_dir.is_dir():
                        yaml_files = list(version_dir.glob("*.yaml"))
                        if yaml_files:
                            key = f"{profile_dir.name}/{version_dir.name}"
                            example_files[key] = yaml_files[0]

    if list_examples or profile is None:
        typer.echo("Available example datasets:")
        typer.echo("")
        for name, path in sorted(example_files.items()):
            typer.echo(f"  {name:30} {path.name}")
        typer.echo("")
        typer.echo("Usage: metaseed example <profile/version> -o output.xlsx")
        typer.echo("       metaseed example miappe/1.1 -o example.xlsx")
        return

    # Handle both "profile/version" and just "profile" (uses latest)
    profile_input = profile.lower()
    if "/" in profile_input:
        example_key = profile_input
    else:
        # Find latest version for profile
        matching = [k for k in example_files if k.startswith(f"{profile_input}/")]
        if not matching:
            echo_error(
                f"No examples for profile '{profile_input}'. Available: {', '.join(sorted(example_files.keys()))}"
            )
            raise typer.Exit(EXIT_CONFIG_ERROR)
        example_key = sorted(matching)[-1]  # Latest version

    if example_key not in example_files:
        echo_error(
            f"Example not found: '{example_key}'. Available: {', '.join(sorted(example_files.keys()))}"
        )
        raise typer.Exit(EXIT_CONFIG_ERROR)

    example_file = example_files[example_key]
    if not example_file.exists():
        echo_error(f"Example file not found: {example_file}")
        raise typer.Exit(EXIT_INPUT_ERROR)

    # Load example data
    try:
        data = yaml.safe_load(example_file.read_text(encoding="utf-8"))
    except yaml.YAMLError as e:
        echo_error(f"Invalid YAML in example: {e}")
        raise typer.Exit(EXIT_INPUT_ERROR) from None

    if output is None:
        # Print to stdout as YAML
        typer.echo(yaml.dump(data, default_flow_style=False, sort_keys=False, allow_unicode=True))
        return

    output_suffix = output.suffix.lower()

    if output_suffix in [".yaml", ".yml"]:
        # Copy YAML file
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(
            yaml.dump(data, default_flow_style=False, sort_keys=False, allow_unicode=True),
            encoding="utf-8",
        )
        echo_success(f"Example exported to {output}")

    elif output_suffix == ".xlsx":
        # Export to Excel
        try:
            _export_example_to_excel(data, output)
            echo_success(f"Example exported to {output}")
        except ImportError:
            echo_error("openpyxl is required for Excel export. Install with: pip install openpyxl")
            raise typer.Exit(EXIT_CONFIG_ERROR) from None

    elif output_suffix == ".json":
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
        echo_success(f"Example exported to {output}")

    else:
        echo_error(f"Unknown output format: {output_suffix}. Use .xlsx, .yaml, or .json")
        raise typer.Exit(EXIT_INPUT_ERROR)
