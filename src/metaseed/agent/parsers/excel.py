"""Excel file parser."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Self

from openpyxl import load_workbook

from metaseed.agent.parsers.registry import ParsedContent, ParsedTable


class ExcelParser:
    """Parser for OOXML Excel files (.xlsx, .xlsm).

    Legacy binary .xls files are not supported: the backing openpyxl library
    only reads the OOXML formats.
    """

    extensions = [".xlsx", ".xlsm"]
    mime_types = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ]

    def can_parse(self: Self, path: Path) -> bool:
        """Check if this parser can handle the file."""
        return path.suffix.lower() in self.extensions

    def parse(self: Self, path: Path) -> ParsedContent:
        """Parse Excel file into structured content.

        Reads all sheets from the workbook. Each sheet becomes a separate
        table in the parsed content.

        Args:
            path: Path to the Excel file.

        Returns:
            ParsedContent with tables for each sheet.
        """
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        tables: list[ParsedTable] = []

        # Capture sheet names before closing the workbook (a read-only workbook
        # releases its backing file on close, so reading them after is unsafe).
        sheet_names = list(workbook.sheetnames)

        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            table = self._parse_sheet(sheet, sheet_name)
            if table:
                tables.append(table)

        workbook.close()

        return ParsedContent(
            source_path=str(path),
            format="excel",
            tables=tables,
            metadata={
                "sheet_count": len(sheet_names),
                "sheet_names": sheet_names,
            },
        )

    def _parse_sheet(self: Self, sheet: Any, sheet_name: str) -> ParsedTable | None:
        """Parse a single sheet into a table.

        Args:
            sheet: Worksheet object.
            sheet_name: Name of the sheet.

        Returns:
            ParsedTable or None if sheet is empty.
        """
        rows_raw: list[list[Any]] = []

        for row in sheet.iter_rows():
            row_values = [self._cell_value(cell) for cell in row]
            # Skip completely empty rows
            if any(v is not None and v != "" for v in row_values):
                rows_raw.append(row_values)

        if not rows_raw:
            return None

        # First non-empty row is headers
        headers = [str(h) if h else f"Column_{i}" for i, h in enumerate(rows_raw[0])]
        data_rows: list[list[Any]] = []

        for row in rows_raw[1:]:
            # Normalize row length to match headers
            if len(row) < len(headers):
                row = row + [None] * (len(headers) - len(row))
            elif len(row) > len(headers):
                row = row[: len(headers)]
            data_rows.append(row)

        return ParsedTable(
            name=sheet_name,
            headers=headers,
            rows=data_rows,
            source_location=f"sheet:{sheet_name}",
        )

    def _cell_value(self: Self, cell: Any) -> Any:
        """Extract value from a cell.

        Args:
            cell: Cell object.

        Returns:
            Cell value, with appropriate type conversion.
        """
        value = cell.value

        if value is None:
            return None

        # Convert datetime to ISO string
        if hasattr(value, "isoformat"):
            return value.isoformat()

        return value
